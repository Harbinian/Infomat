#!/usr/bin/env python3
"""
BizMapper - 业务关系提取工具
从 PNG 业务流程图中提取业务能力、业务流程、应用系统的映射关系
"""

import sys
import os
import json
import base64
import urllib.request
import urllib.error
from pathlib import Path

# MiniMax API 配置
MINIMAX_API_HOST = os.getenv("MINIMAX_API_HOST", "https://api.minimaxi.com")
MINIMAX_API_KEY = os.getenv("MINIMAX_API_KEY")

ANALYSIS_PROMPT = """你是一个业务架构分析专家。请分析这张业务流程图，提取结构化数据。

## 图片结构
- 列1：业务能力（一级分类）
- 列2：业务流程（二级分类）
- 列3：应用系统（IT系统）
- 连线：直线，连接相邻列

## 输出要求
请以 JSON 格式输出，包含以下字段：

1. `capabilities`: 业务能力列表
   - `name`: 一级业务能力名称
   - `l3`: 该能力下的业务流程列表

2. `systems`: 应用系统列表
   - `id`: 系统ID（英文或数字，如 s1, system_1）
   - `name`: 系统名称

3. `connections`: 连线关系列表
   - `capName`: 业务能力名称
   - `procName`: 业务流程名称
   - `sysId`: 应用系统ID

## 注意事项
- 如果某条连线的应用系统不确定，该字段留空
- 同一分组的列（1-2-3-2-1结构中的列1+列4属于业务能力）需合并
- 只提取确定的连线关系，不确定时返回空字符串

请直接输出 JSON，不要有其他内容："""


def call_minimax_vision(image_path: str, prompt: str) -> dict:
    """调用 MiniMax Vision API 分析图片

    Args:
        image_path: 图片文件路径
        prompt: 分析提示词

    Returns:
        API 响应字典，包含 analysis 字段
    """
    if not MINIMAX_API_KEY:
        raise RuntimeError("MINIMAX_API_KEY environment variable not set")

    # 读取图片并转换为 base64
    with open(image_path, "rb") as f:
        img_data = base64.b64encode(f.read()).decode()

    # 检测图片格式
    if image_path.lower().endswith('.png'):
        img_format = 'png'
    elif image_path.lower().endswith(('.jpg', '.jpeg')):
        img_format = 'jpeg'
    elif image_path.lower().endswith('.webp'):
        img_format = 'webp'
    else:
        img_format = 'png'  # 默认使用 PNG

    image_url = f"data:image/{img_format};base64,{img_data}"

    # 调用 MiniMax VLM API
    url = f"{MINIMAX_API_HOST}/v1/coding_plan/vlm"
    headers = {
        "Authorization": f"Bearer {MINIMAX_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "prompt": prompt,
        "image_url": image_url
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")

    try:
        with urllib.request.urlopen(req, timeout=120) as response:
            result = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8") if e.fp else ""
        raise RuntimeError(f"HTTP Error {e.code}: {error_body}")
    except urllib.error.URLError as e:
        raise RuntimeError(f"URL Error: {e.reason}")

    # 检查 API 错误
    base_resp = result.get("base_resp", {})
    if base_resp.get("status_code") != 0:
        raise RuntimeError(f"API Error: {base_resp.get('status_msg')}")

    return {"analysis": result.get("content", "")}


def parse_analysis_result(raw_result: dict) -> dict:
    """解析 API 返回的原始结果"""
    try:
        content = raw_result.get("analysis", "")

        # 尝试提取 JSON
        if isinstance(content, str):
            # 去掉 markdown 代码块标记
            content = content.strip()
            if content.startswith("```json"):
                content = content[7:]
            elif content.startswith("```"):
                content = content[3:]
            if content.endswith("```"):
                content = content[:-3]

            return json.loads(content.strip())

        return content
    except json.JSONDecodeError as e:
        raise RuntimeError(f"解析失败: {e}\n原始内容: {content[:500]}")


def validate_data(data: dict) -> list:
    """验证解析结果，返回错误列表"""
    errors = []

    if "capabilities" not in data:
        errors.append("缺少 capabilities 字段")
    if "systems" not in data:
        errors.append("缺少 systems 字段")
    if "connections" not in data:
        errors.append("缺少 connections 字段")

    # 验证 connections 中的引用完整性
    cap_names = {c["name"] for c in data.get("capabilities", [])}
    all_l3 = set()
    for c in data.get("capabilities", []):
        all_l3.update(c.get("l3", []))
    sys_ids = {s["id"] for s in data.get("systems", [])}

    for i, conn in enumerate(data.get("connections", [])):
        if conn.get("capName") and conn["capName"] not in cap_names:
            errors.append(f"连线 {i}: 未找到业务能力 '{conn['capName']}'")
        if conn.get("procName") and conn["procName"] not in all_l3:
            errors.append(f"连线 {i}: 未找到业务流程 '{conn['procName']}'")
        if conn.get("sysId") and conn["sysId"] not in sys_ids:
            errors.append(f"连线 {i}: 未找到系统 '{conn['sysId']}'")

    return errors


def analyze_image(image_path: str) -> dict:
    """分析图片并返回结构化数据"""
    print(f"正在分析图片: {image_path}")
    raw_result = call_minimax_vision(image_path, ANALYSIS_PROMPT)
    return parse_analysis_result(raw_result)


def test_api_connection():
    """测试 MiniMax API 连接"""
    print("Testing MiniMax API connection...")

    if not MINIMAX_API_KEY:
        print("FAIL: MINIMAX_API_KEY environment variable not set")
        return False

    # 创建一个简单的测试图片（1x1 绿色像素 PNG）
    test_image_base64 = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="

    try:
        url = f"{MINIMAX_API_HOST}/v1/coding_plan/vlm"
        headers = {
            "Authorization": f"Bearer {MINIMAX_API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "prompt": "Describe this image in one word.",
            "image_url": f"data:image/png;base64,{test_image_base64}"
        }

        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")

        with urllib.request.urlopen(req, timeout=120) as response:
            result = json.loads(response.read().decode("utf-8"))

        base_resp = result.get("base_resp", {})

        if base_resp.get("status_code") == 0:
            print(f"SUCCESS: API connection working")
            print(f"Response: {result.get('content', '')}")
            return True
        else:
            print(f"FAIL: API returned error: {base_resp.get('status_msg')}")
            return False

    except Exception as e:
        print(f"FAIL: {str(e)}")
        return False


def generate_excel(data: dict, output_path: str):
    """生成美化 Excel 映射表"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 构建关系列表
    relations = []
    cap_l3_map = {c["name"]: c.get("l3", []) for c in data.get("capabilities", [])}

    for conn in data.get("connections", []):
        cap_name = conn.get("capName", "")
        proc_name = conn.get("procName", "")
        sys_id = conn.get("sysId", "")
        sys_name = ""

        # 查找系统名称
        for s in data.get("systems", []):
            if s["id"] == sys_id:
                sys_name = s["name"]
                break

        # 查找该流程属于哪个一级业务能力
        l1_cap = ""
        for cap, l3_list in cap_l3_map.items():
            if proc_name in l3_list:
                l1_cap = cap
                break

        relations.append({
            "一级业务能力": l1_cap,
            "业务流程": proc_name,
            "应用系统": sys_name
        })

    # 创建 DataFrame
    df = pd.DataFrame(relations)

    # 写入 Excel（基础）
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='映射表')

    # 美化
    style_excel(output_path)

    return relations


def style_excel(output_path: str):
    """美化 Excel 文件"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = load_workbook(output_path)
    ws = wb.active

    # 表头样式
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    # 边框
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 样式应用到表头
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # 一级业务能力配色
    colors = ["D6E4F0", "D5F5E3", "FEF9E7", "FADBD8", "E8DAEF", "D4EFDF"]
    cap_index = 0

    # 按一级业务能力分组并应用颜色
    current_cap = None
    cap_start_row = 2

    for row in range(2, ws.max_row + 1):
        cap = ws.cell(row=row, column=1).value
        if cap != current_cap:
            if current_cap is not None and row > cap_start_row:
                # 应用颜色到之前的组
                color = colors[cap_index % len(colors)]
                for r in range(cap_start_row, row):
                    for c in range(1, 4):
                        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)

                cap_index += 1

            current_cap = cap
            cap_start_row = row

    # 最后一组
    if current_cap is not None:
        color = colors[cap_index % len(colors)]
        for r in range(cap_start_row, ws.max_row + 1):
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)

    # 应用边框和对齐到所有数据单元格
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 合并一级业务能力中的相同单元格
    merge_same_cells(ws, 1, 2, ws.max_row)

    # 列宽
    col_widths = {"A": 22, "B": 28, "C": 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 行高
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    wb.save(output_path)


def merge_same_cells(ws, col_index, start_row, end_row):
    """合并指定列中连续相同值的单元格"""
    merge_start = start_row
    for row in range(start_row + 1, end_row + 2):
        current = ws.cell(row=row, column=col_index).value
        prev = ws.cell(row=row - 1, column=col_index).value
        if current != prev or row == end_row + 2:
            if row - 1 > merge_start:
                ws.merge_cells(
                    start_row=merge_start, start_column=col_index,
                    end_row=row - 1, end_column=col_index
                )
            merge_start = row


def excel_to_json(excel_path: str, output_path: str = None):
    """读取审核后的 Excel，转换为 Infomat JSON"""
    import pandas as pd

    # 读取 Excel
    df = pd.read_excel(excel_path, sheet_name='映射表')

    # 构建数据结构
    capabilities_map = {}  # name -> {name, l3: []}
    systems_map = {}       # id -> {id, name}
    connections = []

    sys_index = 1
    for _, row in df.iterrows():
        cap_name = row["一级业务能力"]
        proc_name = row["业务流程"]
        sys_name = row["应用系统"]

        # 更新 capabilities
        if cap_name not in capabilities_map:
            capabilities_map[cap_name] = {"name": cap_name, "l3": []}

        # 添加 l3（业务流程）
        if proc_name and proc_name not in capabilities_map[cap_name]["l3"]:
            capabilities_map[cap_name]["l3"].append(proc_name)

        # 处理系统（如果名称为空则跳过）
        if sys_name and pd.notna(sys_name):
            # 生成简单的 sysId
            sys_id = f"s{sys_index}"
            for existing in systems_map.values():
                if existing["name"] == sys_name:
                    sys_id = existing["id"]
                    break
            else:
                systems_map[sys_id] = {"id": sys_id, "name": sys_name}
                sys_index += 1

            # 添加连线
            connections.append({
                "capName": cap_name,
                "procName": proc_name,
                "sysId": sys_id
            })

    # 构建最终结构
    result = {
        "capabilities": list(capabilities_map.values()),
        "systems": list(systems_map.values()),
        "connections": connections
    }

    # 保存
    if output_path is None:
        import datetime
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        output_path = f"output/infomat_data_{date_str}.json"

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"JSON 已保存至: {output_path}")
    return result


def main():
    if len(sys.argv) < 2:
        print("BizMapper v1.0 - 业务关系提取工具")
        print("")
        print("用法:")
        print("  生成 Excel: python bizmapper.py <图片路径> [输出目录]")
        print("  Excel转JSON: python bizmapper.py --to-json <Excel路径> [输出目录]")
        print("")
        print("示例:")
        print("  python bizmapper.py sample-diagram.png")
        print("  python bizmapper.py --to-json output/映射表_20260429.xlsx")
        return

    output_dir = "output"
    if len(sys.argv) >= 3:
        output_dir = sys.argv[2]

    if sys.argv[1] == "--to-json":
        # Excel 转 JSON 模式
        excel_path = sys.argv[2]
        output_dir = sys.argv[3] if len(sys.argv) >= 4 else "output"
        os.makedirs(output_dir, exist_ok=True)
        import datetime
        json_path = os.path.join(output_dir, f"infomat_data_{datetime.datetime.now().strftime('%Y%m%d')}.json")
        excel_to_json(excel_path, json_path)
    else:
        # 分析图片生成 Excel 模式
        image_path = sys.argv[1]
        os.makedirs(output_dir, exist_ok=True)

        # 分析
        data = analyze_image(image_path)

        # 验证
        errors = validate_data(data)
        if errors:
            print("警告 - 数据验证发现以下问题:")
            for err in errors:
                print(f"  - {err}")
            print("")

        # 生成 Excel
        import datetime
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        excel_path = os.path.join(output_dir, f"映射表_{date_str}.xlsx")

        print(f"正在生成 Excel: {excel_path}")
        generate_excel(data, excel_path)
        print(f"Excel 已保存至: {excel_path}")
        print("")
        print("请审核 Excel 中的映射关系，调整后运行以下命令转换为 JSON:")
        print(f"  python bizmapper.py --to-json {excel_path}")


if __name__ == "__main__":
    main()
