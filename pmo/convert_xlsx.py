import csv
import json
import pathlib
import re
from datetime import date, datetime

import openpyxl


ROOT = pathlib.Path(__file__).resolve().parent


def norm_text(v):
    if v is None:
        return ""
    return str(v).strip()


def norm_int(v):
    s = norm_text(v)
    try:
        return int(float(s))
    except Exception:
        return 0


def norm_date(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()

    s = norm_text(v)
    if not s:
        return ""

    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.match(r"^(\d{4})年(\d{1,2})月(\d{1,2})日$", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    return ""


def read_tasks_from_xlsx(xlsx_path: pathlib.Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    target = None
    for sn in wb.sheetnames:
        if "Project" in sn and "导入任务" in sn:
            target = sn
            break
    if target is None:
        raise RuntimeError(f"Sheet containing 'Project导入任务表' not found. Available: {wb.sheetnames}")
    ws = wb[target]

    header = [norm_text(c.value) for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    def get(row, name):
        i = idx.get(name)
        if i is None:
            return None
        if i >= len(row):
            return None
        return row[i]

    tasks = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        task_name = norm_text(get(r, "任务名称"))
        if not task_name:
            continue

        task = {
            "id": norm_int(get(r, "ID")),
            "wbs": norm_text(get(r, "WBS")),
            "name": task_name,
            "type": norm_text(get(r, "任务类型")),
            "duration": norm_text(get(r, "工期")),
            "start": norm_date(get(r, "开始时间")),
            "finish": norm_date(get(r, "完成时间")),
            "predecessors": norm_text(get(r, "前置任务")),
            "resources": norm_text(get(r, "主责资源")) or norm_text(get(r, "资源名称")),
            "department": norm_text(get(r, "责任部门")),
            "vendor": norm_text(get(r, "供应商")),
            "reviewer": norm_text(get(r, "审核人/审批组")),
            "risk": norm_text(get(r, "风险等级")),
            "milestone": norm_text(get(r, "里程碑")),
            "deliverable": norm_text(get(r, "交付物")),
            "notes": norm_text(get(r, "备注")),
        }
        tasks.append(task)

    ids = set()
    dup = set()
    for t in tasks:
        if not t["id"]:
            continue
        if t["id"] in ids:
            dup.add(t["id"])
        ids.add(t["id"])
    if dup:
        raise RuntimeError(f"Duplicate task id(s): {', '.join(str(x) for x in sorted(dup))}")

    return tasks


def write_tasks_json(tasks, out_path: pathlib.Path):
    out_path.write_text(json.dumps(tasks, ensure_ascii=False, indent=2), encoding="utf-8")


def write_tasks_csv(tasks, out_path: pathlib.Path):
    header = [
        "ID",
        "WBS",
        "任务名称",
        "任务类型",
        "工期",
        "开始时间",
        "完成时间",
        "前置任务",
        "资源名称",
        "责任部门",
        "供应商",
        "审核人/审批组",
        "风险等级",
        "里程碑",
        "交付物",
        "备注",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for t in tasks:
            w.writerow(
                [
                    t.get("id") or "",
                    t.get("wbs") or "",
                    t.get("name") or "",
                    t.get("type") or "",
                    t.get("duration") or "",
                    t.get("start") or "",
                    t.get("finish") or "",
                    t.get("predecessors") or "",
                    t.get("resources") or "",
                    t.get("department") or "",
                    t.get("vendor") or "",
                    t.get("reviewer") or "",
                    t.get("risk") or "",
                    t.get("milestone") or "",
                    t.get("deliverable") or "",
                    t.get("notes") or "",
                ]
            )


def main():
    xlsx_path = ROOT / "信息化项目_Project_H5可用.xlsx"
    tasks_path = ROOT / "tasks.json"
    csv_path = ROOT / "信息化项目.csv"
    react_tasks_path = ROOT / "gantt-react" / "public" / "tasks.json"

    tasks = read_tasks_from_xlsx(xlsx_path)

    write_tasks_json(tasks, tasks_path)
    write_tasks_csv(tasks, csv_path)
    react_tasks_path.write_text(tasks_path.read_text(encoding="utf-8"), encoding="utf-8")

    print(f"Wrote {len(tasks)} tasks")


if __name__ == "__main__":
    main()

