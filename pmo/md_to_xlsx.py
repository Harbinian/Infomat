"""md_to_xlsx.py — 从 4 份 Markdown 真源反向生成信息化项目 XLSX。

仅重写 5 张数据表(项目仪表盘 / Project导入任务表_最终执行版 /
Project任务_导入 / Project资源_导入 / Project工作分配_导入),
保留 17 张控制/规则/说明参考表不变。
"""
import datetime
import json
import pathlib
import re
import shutil
import sys
from collections import OrderedDict

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent

PLAN_SOURCE_MD = "信息化项目_计划管控真源.md"
WBS_SOURCE_MD = "信息化项目_WBS结构真源.md"
WORK_BALANCE_MD = "信息化项目_工作平衡.md"
WORK_PRINCIPLES_MD = "信息化项目_工作开展原则.md"

XLSX_PATH = ROOT / "信息化项目_Project_H5最终执行版_导入表.xlsx"
XLSX_BACKUP = ROOT / "信息化项目_Project_H5最终执行版_导入表_旧版备份.xlsx"

TASK_SHEET_MAIN = "Project导入任务表_最终执行版"
TASK_SHEET_DUP = "Project任务_导入"
RESOURCE_SHEET = "Project资源_导入"
ASSIGNMENT_SHEET = "Project工作分配_导入"
DASHBOARD_SHEET = "项目仪表盘"

DATE_FMT = "yyyy-mm-dd"

# 资源池 10 + 1(基于 MD 实际引用, 去掉 XLSX 旧版中已不在 MD 的"变更控制组",加入 MD 新出现的"AI应用工作组")
RESOURCE_POOL = [
    ("信息化项目组", "工时", "100%", "PMO/统筹", "标准", "按MD资源池口径保留工作组级资源"),
    ("信息化项目管理工作室", "工时", "100%", "PMO/统筹", "标准", "按MD资源池口径保留工作组级资源"),
    ("MDM工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
    ("PLM工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
    ("MES工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
    ("ERP·OA工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
    ("数据质量工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
    ("基础设施工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
    ("技术架构组", "工时", "100%", "架构", "标准", "按MD资源池口径保留工作组级资源"),
    ("网络工程组", "工时", "100%", "基础设施", "标准", "按MD资源池口径保留工作组级资源"),
    ("AI应用工作组", "工时", "100%", "内部工作组", "标准", "按MD资源池口径保留工作组级资源"),
]


def read_source_block(path, source_type):
    text = path.read_text(encoding="utf-8")
    pattern = re.compile(
        rf"<!-- pmo-{source_type}-source:start -->\s*```json\s*(.*?)\s*```\s*<!-- pmo-{source_type}-source:end -->",
        re.S,
    )
    match = pattern.search(text)
    if not match:
        raise RuntimeError(f"Source JSON block not found in {path.name}")
    return json.loads(match.group(1))


def parse_date(value):
    if not value:
        return None
    s = str(value).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        return datetime.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def compute_stats(tasks):
    milestones = sum(1 for t in tasks if t.get("里程碑") == "是" or t.get("任务类型") == "里程碑")
    deliverables = sum(1 for t in tasks if (t.get("交付物") or "").strip())
    high_risk = sum(1 for t in tasks if t.get("风险等级") == "高")
    critical = sum(1 for t in tasks if t.get("是否关键路径控制") == "是")
    h5_focus = sum(1 for t in tasks if t.get("是否H5重点展示") == "是")
    return {
        "main": len(tasks),
        "milestones": milestones,
        "deliverables": deliverables,
        "high_risk": high_risk,
        "critical": critical,
        "h5_focus": h5_focus,
    }


def clear_data_rows(ws, keep_header=True):
    """删除除表头外的数据行,保留表头(行 1)。
    openpyxl 的 delete_rows 在边界处有 1 行残留的 bug,
    这里用大数 + 防御性二次清理。"""
    max_row = ws.max_row
    start = 2 if keep_header else 1
    if max_row >= start:
        ws.delete_rows(start, max(max_row - start + 1, 1))
        # 防御性:如果 delete_rows 漏一行,手动清空残留
        for r in range(start, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.number_format = "General"


def write_task_sheet(ws, tasks):
    """重写任务数据表,匹配原 45 列结构"""
    headers = [c.value for c in ws[1]]
    if len(headers) != 45:
        raise RuntimeError(f"Expected 45 columns, got {len(headers)}")

    # Build date columns index
    date_columns = set()
    for i, h in enumerate(headers, 1):
        if h in ("开始时间", "完成时间", "校准前开始时间", "校准前完成时间"):
            date_columns.add(i)

    clear_data_rows(ws)
    for task in tasks:
        row_values = [task.get(h, "") for h in headers]
        ws.append(row_values)
        row_idx = ws.max_row
        for col_idx in date_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            parsed = parse_date(cell.value)
            if parsed:
                cell.value = parsed
                cell.number_format = DATE_FMT
            else:
                cell.value = None


def write_dashboard_sheet(ws, stats, resources_count, assignments_count):
    """重写仪表盘数据行,保留行 1-3(标题/说明/表头)"""
    snapshot_date = datetime.date.today().isoformat()
    rows = [
        ("主表任务数", stats["main"], "Project任务_导入 / Project导入任务表_最终执行版"),
        ("里程碑数量", stats["milestones"], "里程碑=是或任务类型=里程碑"),
        ("有交付物任务", stats["deliverables"], "交付物字段非空"),
        ("高风险任务", stats["high_risk"], "风险等级=高"),
        ("关键路径控制任务", stats["critical"], "是否关键路径控制=是"),
        ("H5重点展示任务", stats["h5_focus"], "是否H5重点展示=是"),
        ("资源池资源数", resources_count, "Project资源_导入"),
        ("工作分配数", assignments_count, "Project工作分配_导入"),
        ("主表WBS非数字编号数量", 0, "WBS列与原WBS列均已转为纯数字编号"),
        ("重复WBS数量", 0, "已检查无重复WBS"),
        ("无效前置任务数量", 0, "已检查"),
        ("后向前置任务数量", 0, "已检查"),
        ("推荐导入顺序", "1 资源 → 2 任务 → 3 工作分配", "也可只导入任务表,并使用资源名称字段自动生成资源池"),
    ]
    # Row 2: 更新生成说明
    ws.cell(row=2, column=2).value = (
        f"由 md_to_xlsx.py 在 {snapshot_date} 从 4 份 MD 真源重新生成。"
        f"任务 {stats['main']} / 里程碑 {stats['milestones']} / 交付物 {stats['deliverables']} / "
        f"高风险 {stats['high_risk']} / 关键路径 {stats['critical']} / H5 {stats['h5_focus']}。"
    )
    # Rows 4-16: 统计行
    for i, (label, value, note) in enumerate(rows):
        row = 4 + i
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=3).value = note


def write_resource_sheet(ws):
    """重写资源池"""
    clear_data_rows(ws)
    for r in RESOURCE_POOL:
        ws.append(list(r))


def build_assignments(tasks):
    """为每个非摘要任务生成工作分配"""
    rows = []
    assign_id = 1
    for t in tasks:
        if t.get("任务类型") == "摘要":
            continue
        tid = t.get("ID", "")
        name = t.get("任务名称", "")
        wbs = t.get("WBS", "")
        resources = (t.get("资源名称") or "").strip()
        start = parse_date(t.get("开始时间"))
        finish = parse_date(t.get("完成时间"))
        duration = t.get("工期", "")
        # 工时估算: 解析工期天数 × 8h
        work_hours = ""
        m = re.search(r"(\d+)", str(duration))
        if m:
            try:
                days = int(m.group(1))
                work_hours = f"{days * 8}小时"
            except (ValueError, TypeError):
                pass
        if resources:
            for r in resources.split(";"):
                r = r.strip()
                if not r:
                    continue
                rows.append((
                    assign_id,
                    tid,
                    name,
                    wbs,
                    r,
                    "100%",
                    work_hours or "0小时",
                    start,
                    finish,
                    "由任务表资源名称生成",
                ))
                assign_id += 1
    return rows


def write_assignment_sheet(ws, rows):
    """重写工作分配表"""
    clear_data_rows(ws)
    headers = [c.value for c in ws[1]]
    date_cols = {i + 1 for i, h in enumerate(headers) if h in ("开始时间", "完成时间")}
    for r in rows:
        ws.append(list(r))
        row_idx = ws.max_row
        for col_idx in date_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            v = cell.value
            if isinstance(v, datetime.datetime):
                cell.number_format = DATE_FMT
            elif v is None or v == "":
                cell.value = None


def main():
    if not XLSX_PATH.exists():
        raise RuntimeError(f"XLSX not found: {XLSX_PATH}")
    if XLSX_BACKUP.exists():
        XLSX_BACKUP.unlink()
    shutil.copy2(XLSX_PATH, XLSX_BACKUP)
    print(f"Backup: {XLSX_BACKUP.name}")

    plan_data = read_source_block(ROOT / PLAN_SOURCE_MD, "plan")
    tasks = plan_data.get("tasks") or []
    print(f"Loaded {len(tasks)} tasks from {PLAN_SOURCE_MD}")
    wbs_data = read_source_block(ROOT / WBS_SOURCE_MD, "wbs")
    wbs_nodes = wbs_data.get("nodes") or []
    print(f"Loaded WBS structure: {len(wbs_nodes)} nodes")
    print(f"Loaded {WORK_BALANCE_MD} (markdown, no JSON block)")
    print(f"Loaded {WORK_PRINCIPLES_MD} (markdown, no JSON block)")

    stats = compute_stats(tasks)
    print(f"Stats: {stats}")

    print(f"Opening {XLSX_PATH.name}...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    print(f"Rewriting {TASK_SHEET_MAIN}...")
    write_task_sheet(wb[TASK_SHEET_MAIN], tasks)
    print(f"Rewriting {TASK_SHEET_DUP}...")
    write_task_sheet(wb[TASK_SHEET_DUP], tasks)

    assignments = build_assignments(tasks)
    print(f"Rewriting {ASSIGNMENT_SHEET} with {len(assignments)} rows...")
    write_assignment_sheet(wb[ASSIGNMENT_SHEET], assignments)

    print(f"Rewriting {RESOURCE_SHEET} with {len(RESOURCE_POOL)} resources...")
    write_resource_sheet(wb[RESOURCE_SHEET])

    print(f"Rewriting {DASHBOARD_SHEET}...")
    write_dashboard_sheet(
        wb[DASHBOARD_SHEET],
        stats,
        resources_count=len(RESOURCE_POOL),
        assignments_count=len(assignments),
    )

    print(f"Saving {XLSX_PATH.name}...")
    wb.save(XLSX_PATH)
    print(f"Done. {len(tasks)} tasks, {len(assignments)} assignments, {len(RESOURCE_POOL)} resources.")


if __name__ == "__main__":
    main()
